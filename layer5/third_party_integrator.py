"""
Third-Party Integrator for DAP System
Layer 5: External Integration & API Services

Provides comprehensive integration with external audit tools, financial systems,
and enterprise applications through standardized connectors and adapters.
"""

import asyncio
import json
import time
import uuid
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Callable, Union, Tuple
import logging
from contextlib import asynccontextmanager
import ssl
import base64
from dataclasses import dataclass, asdict
from urllib.parse import urljoin, urlencode
import xml.etree.ElementTree as ET

try:
    import aiohttp
    import aiofiles
    AIOHTTP_AVAILABLE = True
except ImportError:
    AIOHTTP_AVAILABLE = False

try:
    import sqlalchemy as sa
    from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
    from sqlalchemy.orm import sessionmaker
    SQLALCHEMY_AVAILABLE = True
except ImportError:
    SQLALCHEMY_AVAILABLE = False

try:
    import ftplib
    import paramiko
    FTP_SSH_AVAILABLE = True
except ImportError:
    FTP_SSH_AVAILABLE = False

try:
    import ldap3
    LDAP_AVAILABLE = True
except ImportError:
    LDAP_AVAILABLE = False

try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

@dataclass
class IntegrationConfig:
    """Configuration for third-party integration"""
    name: str
    type: str  # 'api', 'database', 'file', 'ftp', 'ldap', etc.
    endpoint: str
    credentials: Dict[str, Any]
    settings: Dict[str, Any]
    timeout: int = 30
    retry_count: int = 3
    enabled: bool = True

@dataclass
class DataMapping:
    """Data mapping configuration for external systems"""
    source_field: str
    target_field: str
    transformation: Optional[str] = None
    validation: Optional[str] = None
    default_value: Optional[Any] = None

@dataclass
class SyncStatus:
    """Synchronization status information"""
    integration_name: str
    last_sync: Optional[str]
    next_sync: Optional[str]
    status: str  # 'success', 'failed', 'in_progress', 'disabled'
    records_synced: int
    errors: List[str]
    duration: float

class ThirdPartyIntegrator:
    """Third-party integration manager for external systems"""

    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        self.setup_logging()

        # Integration registry
        self.integrations = {}
        self.data_mappings = {}
        self.sync_schedules = {}

        # Connection pools
        self.http_session = None
        self.db_engines = {}
        self.ftp_connections = {}

        # Status tracking
        self.sync_statuses = {}
        self.active_syncs = set()

        # Background tasks
        self.scheduler_running = False
        self.sync_tasks = []

        self.initialize_integrator()

    def setup_logging(self):
        """Setup enhanced logging for third-party integrator"""
        self.logger = logging.getLogger(f"{__name__}.ThirdPartyIntegrator")

        if not self.logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )
            handler.setFormatter(formatter)
            self.logger.addHandler(handler)
            self.logger.setLevel(logging.INFO)

    def initialize_integrator(self):
        """Initialize the third-party integrator"""
        self.setup_http_session()
        self.load_integrations()
        self.setup_default_mappings()

    def setup_http_session(self):
        """Setup HTTP session for API integrations"""
        if AIOHTTP_AVAILABLE:
            timeout = aiohttp.ClientTimeout(total=30)
            connector = aiohttp.TCPConnector(
                limit=100,
                limit_per_host=10,
                ssl=ssl.create_default_context()
            )
            self.http_session = aiohttp.ClientSession(
                timeout=timeout,
                connector=connector
            )

    def load_integrations(self):
        """Load integration configurations"""
        # Default integration configurations
        default_integrations = {
            'kingdee_k3': IntegrationConfig(
                name='金蝶K3系统',
                type='database',
                endpoint='mssql://server/database',
                credentials={'username': '', 'password': ''},
                settings={
                    'tables': ['gl_voucher', 'gl_balance', 'ar_bill', 'ap_bill'],
                    'encoding': 'utf-8',
                    'date_format': '%Y-%m-%d'
                }
            ),
            'ufida_u8': IntegrationConfig(
                name='用友U8系统',
                type='database',
                endpoint='mssql://server/database',
                credentials={'username': '', 'password': ''},
                settings={
                    'tables': ['gl_accvouch', 'gl_accsum', 'ar_djz', 'ap_djz'],
                    'encoding': 'gb2312',
                    'date_format': '%Y-%m-%d'
                }
            ),
            'sap_erp': IntegrationConfig(
                name='SAP ERP系统',
                type='api',
                endpoint='https://sap-server:8000/sap/bc/rest/',
                credentials={'username': '', 'password': '', 'client': ''},
                settings={
                    'auth_type': 'basic',
                    'modules': ['FI', 'CO', 'MM', 'SD'],
                    'date_format': 'YYYYMMDD'
                }
            ),
            'excel_reports': IntegrationConfig(
                name='Excel报表',
                type='file',
                endpoint='/shared/reports/',
                credentials={},
                settings={
                    'file_patterns': ['*.xlsx', '*.xls'],
                    'auto_detect_sheets': True,
                    'header_row': 1
                }
            ),
            'ftp_server': IntegrationConfig(
                name='FTP文件服务器',
                type='ftp',
                endpoint='ftp://ftp.company.com/',
                credentials={'username': '', 'password': ''},
                settings={
                    'passive_mode': True,
                    'binary_mode': True,
                    'download_path': '/downloads/'
                }
            ),
            'ldap_directory': IntegrationConfig(
                name='LDAP目录服务',
                type='ldap',
                endpoint='ldap://ldap.company.com:389',
                credentials={'username': '', 'password': ''},
                settings={
                    'base_dn': 'dc=company,dc=com',
                    'user_filter': '(objectClass=person)',
                    'group_filter': '(objectClass=group)'
                }
            ),
            'audit_software': IntegrationConfig(
                name='审计软件接口',
                type='api',
                endpoint='https://audit-software.com/api/v1/',
                credentials={'api_key': '', 'secret': ''},
                settings={
                    'auth_type': 'api_key',
                    'export_formats': ['json', 'xml', 'csv'],
                    'batch_size': 1000
                }
            )
        }

        # Load from config
        configured_integrations = self.config.get('integrations', {})

        # Merge configurations
        for name, config_dict in configured_integrations.items():
            if isinstance(config_dict, dict):
                integration = IntegrationConfig(**config_dict)
                self.integrations[name] = integration

        # Add defaults for missing integrations
        for name, integration in default_integrations.items():
            if name not in self.integrations:
                self.integrations[name] = integration

        self.logger.info(f"Loaded {len(self.integrations)} integration configurations")

    def setup_default_mappings(self):
        """Setup default data mappings for common systems"""
        # 金蝶K3映射
        self.data_mappings['kingdee_k3'] = {
            'voucher': [
                DataMapping('FNumber', 'voucher_number'),
                DataMapping('FDate', 'voucher_date'),
                DataMapping('FExplanation', 'description'),
                DataMapping('FDebit', 'debit_amount'),
                DataMapping('FCredit', 'credit_amount'),
                DataMapping('FAccountNumber', 'account_code')
            ],
            'balance': [
                DataMapping('FAccountNumber', 'account_code'),
                DataMapping('FAccountName', 'account_name'),
                DataMapping('FBeginBalance', 'beginning_balance'),
                DataMapping('FEndBalance', 'ending_balance'),
                DataMapping('FDebitAmount', 'debit_amount'),
                DataMapping('FCreditAmount', 'credit_amount')
            ]
        }

        # 用友U8映射
        self.data_mappings['ufida_u8'] = {
            'voucher': [
                DataMapping('ino_id', 'voucher_number'),
                DataMapping('dbill_date', 'voucher_date'),
                DataMapping('cexch_name', 'description'),
                DataMapping('md', 'debit_amount'),
                DataMapping('mc', 'credit_amount'),
                DataMapping('ccode', 'account_code')
            ],
            'balance': [
                DataMapping('ccode', 'account_code'),
                DataMapping('ccode_name', 'account_name'),
                DataMapping('mb', 'beginning_balance'),
                DataMapping('me', 'ending_balance'),
                DataMapping('md', 'debit_amount'),
                DataMapping('mc', 'credit_amount')
            ]
        }

        # SAP映射
        self.data_mappings['sap_erp'] = {
            'journal': [
                DataMapping('BELNR', 'document_number'),
                DataMapping('GJAHR', 'fiscal_year'),
                DataMapping('BLDAT', 'document_date'),
                DataMapping('BKTXT', 'header_text'),
                DataMapping('DMBTR', 'amount_local'),
                DataMapping('SAKNR', 'gl_account')
            ],
            'balance': [
                DataMapping('SAKNR', 'gl_account'),
                DataMapping('TXT50', 'account_text'),
                DataMapping('HSL01', 'period_01_balance'),
                DataMapping('HSL12', 'period_12_balance'),
                DataMapping('UMSOL', 'debit_credit_indicator')
            ]
        }

    async def start_integrator(self):
        """Start the third-party integrator"""
        self.logger.info("Starting Third-Party Integrator...")

        # Start scheduler
        if not self.scheduler_running:
            self.scheduler_running = True
            asyncio.create_task(self.sync_scheduler())

    async def stop_integrator(self):
        """Stop the third-party integrator"""
        self.logger.info("Stopping Third-Party Integrator...")

        self.scheduler_running = False

        # Close HTTP session
        if self.http_session:
            await self.http_session.close()

        # Close database connections
        for engine in self.db_engines.values():
            await engine.dispose()

        # Close FTP connections
        for ftp in self.ftp_connections.values():
            try:
                ftp.quit()
            except:
                pass

    async def sync_scheduler(self):
        """Background scheduler for automatic synchronization"""
        while self.scheduler_running:
            try:
                current_time = datetime.now()

                for integration_name, schedule in self.sync_schedules.items():
                    if not schedule.get('enabled', True):
                        continue

                    last_sync = schedule.get('last_sync')
                    interval = schedule.get('interval', 3600)  # Default 1 hour

                    # Check if sync is due
                    if not last_sync or (current_time - last_sync).total_seconds() >= interval:
                        if integration_name not in self.active_syncs:
                            asyncio.create_task(self.sync_integration(integration_name))

                await asyncio.sleep(60)  # Check every minute

            except Exception as e:
                self.logger.error(f"Scheduler error: {e}")
                await asyncio.sleep(60)

    async def sync_integration(self, integration_name: str, sync_config: Dict[str, Any] = None) -> SyncStatus:
        """Synchronize data with specific integration"""
        start_time = time.time()
        self.active_syncs.add(integration_name)

        try:
            self.logger.info(f"Starting sync for {integration_name}")

            integration = self.integrations.get(integration_name)
            if not integration or not integration.enabled:
                raise ValueError(f"Integration {integration_name} not found or disabled")

            # Update status
            sync_status = SyncStatus(
                integration_name=integration_name,
                last_sync=None,
                next_sync=None,
                status='in_progress',
                records_synced=0,
                errors=[],
                duration=0
            )
            self.sync_statuses[integration_name] = sync_status

            # Route to appropriate sync method
            if integration.type == 'database':
                result = await self.sync_database_integration(integration, sync_config)
            elif integration.type == 'api':
                result = await self.sync_api_integration(integration, sync_config)
            elif integration.type == 'file':
                result = await self.sync_file_integration(integration, sync_config)
            elif integration.type == 'ftp':
                result = await self.sync_ftp_integration(integration, sync_config)
            elif integration.type == 'ldap':
                result = await self.sync_ldap_integration(integration, sync_config)
            else:
                raise ValueError(f"Unsupported integration type: {integration.type}")

            # Update status
            sync_status.status = 'success'
            sync_status.records_synced = result.get('records_synced', 0)
            sync_status.last_sync = datetime.now().isoformat()
            sync_status.duration = time.time() - start_time

            # Update schedule
            if integration_name in self.sync_schedules:
                self.sync_schedules[integration_name]['last_sync'] = datetime.now()

            self.logger.info(f"Sync completed for {integration_name}: {sync_status.records_synced} records")
            return sync_status

        except Exception as e:
            self.logger.error(f"Sync failed for {integration_name}: {e}")

            sync_status = self.sync_statuses.get(integration_name, SyncStatus(
                integration_name=integration_name,
                last_sync=None,
                next_sync=None,
                status='failed',
                records_synced=0,
                errors=[],
                duration=0
            ))

            sync_status.status = 'failed'
            sync_status.errors.append(str(e))
            sync_status.duration = time.time() - start_time

            return sync_status

        finally:
            self.active_syncs.discard(integration_name)

    async def sync_database_integration(self, integration: IntegrationConfig, sync_config: Dict[str, Any] = None) -> Dict[str, Any]:
        """Synchronize with database integration"""
        if not SQLALCHEMY_AVAILABLE:
            raise RuntimeError("SQLAlchemy not available for database integration")

        records_synced = 0

        try:
            # Create database engine
            engine_key = f"{integration.name}_engine"
            if engine_key not in self.db_engines:
                connection_string = self.build_connection_string(integration)
                self.db_engines[engine_key] = create_async_engine(connection_string)

            engine = self.db_engines[engine_key]

            # Get tables to sync
            tables = integration.settings.get('tables', [])
            mappings = self.data_mappings.get(integration.name.replace('系统', '').lower(), {})

            async with engine.begin() as conn:
                for table_name in tables:
                    try:
                        # Build query
                        query = f"SELECT * FROM {table_name}"

                        # Add date filter if configured
                        date_filter = sync_config.get('date_filter') if sync_config else None
                        if date_filter:
                            query += f" WHERE {date_filter}"

                        # Add limit
                        limit = sync_config.get('limit', 1000) if sync_config else 1000
                        query += f" LIMIT {limit}" if 'mysql' in integration.endpoint.lower() else f" TOP {limit}"

                        # Execute query
                        result = await conn.execute(sa.text(query))
                        rows = result.fetchall()

                        # Apply data mappings
                        table_mappings = mappings.get(table_name.split('_')[-1], [])
                        mapped_data = []

                        for row in rows:
                            mapped_row = {}
                            row_dict = dict(row._mapping) if hasattr(row, '_mapping') else dict(row)

                            for mapping in table_mappings:
                                if mapping.source_field in row_dict:
                                    value = row_dict[mapping.source_field]

                                    # Apply transformation
                                    if mapping.transformation:
                                        value = self.apply_transformation(value, mapping.transformation)

                                    mapped_row[mapping.target_field] = value
                                elif mapping.default_value is not None:
                                    mapped_row[mapping.target_field] = mapping.default_value

                            mapped_data.append(mapped_row)

                        # Store or process mapped data
                        await self.process_synced_data(integration.name, table_name, mapped_data)

                        records_synced += len(rows)

                    except Exception as e:
                        self.logger.error(f"Error syncing table {table_name}: {e}")

            return {'records_synced': records_synced}

        except Exception as e:
            self.logger.error(f"Database sync error: {e}")
            raise

    async def sync_api_integration(self, integration: IntegrationConfig, sync_config: Dict[str, Any] = None) -> Dict[str, Any]:
        """Synchronize with API integration"""
        if not AIOHTTP_AVAILABLE:
            raise RuntimeError("aiohttp not available for API integration")

        records_synced = 0

        try:
            # Setup authentication
            auth_headers = await self.setup_api_auth(integration)

            # Get endpoints to sync
            endpoints = sync_config.get('endpoints', []) if sync_config else []
            if not endpoints:
                # Default endpoints based on system type
                if 'sap' in integration.name.lower():
                    endpoints = ['/fi/documents', '/fi/balances']
                elif 'audit' in integration.name.lower():
                    endpoints = ['/reports', '/audit-trails']

            for endpoint in endpoints:
                try:
                    # Build request URL
                    url = urljoin(integration.endpoint, endpoint)

                    # Add query parameters
                    params = sync_config.get('params', {}) if sync_config else {}

                    async with self.http_session.get(url, headers=auth_headers, params=params) as response:
                        if response.status == 200:
                            # Parse response
                            content_type = response.headers.get('content-type', '')

                            if 'json' in content_type:
                                data = await response.json()
                            elif 'xml' in content_type:
                                text = await response.text()
                                data = self.parse_xml_data(text)
                            else:
                                text = await response.text()
                                data = {'raw_data': text}

                            # Apply data mappings
                            mappings = self.data_mappings.get(integration.name.lower(), {})
                            endpoint_name = endpoint.strip('/').split('/')[-1]
                            endpoint_mappings = mappings.get(endpoint_name, [])

                            mapped_data = []
                            records = data if isinstance(data, list) else data.get('records', [data])

                            for record in records:
                                mapped_record = {}
                                for mapping in endpoint_mappings:
                                    if mapping.source_field in record:
                                        value = record[mapping.source_field]

                                        # Apply transformation
                                        if mapping.transformation:
                                            value = self.apply_transformation(value, mapping.transformation)

                                        mapped_record[mapping.target_field] = value
                                    elif mapping.default_value is not None:
                                        mapped_record[mapping.target_field] = mapping.default_value

                                mapped_data.append(mapped_record)

                            # Store or process mapped data
                            await self.process_synced_data(integration.name, endpoint_name, mapped_data)

                            records_synced += len(records)

                        else:
                            self.logger.error(f"API request failed: {response.status} - {await response.text()}")

                except Exception as e:
                    self.logger.error(f"Error syncing endpoint {endpoint}: {e}")

            return {'records_synced': records_synced}

        except Exception as e:
            self.logger.error(f"API sync error: {e}")
            raise

    async def sync_file_integration(self, integration: IntegrationConfig, sync_config: Dict[str, Any] = None) -> Dict[str, Any]:
        """Synchronize with file integration"""
        records_synced = 0

        try:
            import os
            import glob

            # Get file path
            file_path = integration.endpoint
            file_patterns = integration.settings.get('file_patterns', ['*.xlsx', '*.csv'])

            # Find files to process
            files_to_process = []
            for pattern in file_patterns:
                pattern_path = os.path.join(file_path, pattern)
                files_to_process.extend(glob.glob(pattern_path))

            # Apply filters
            if sync_config and sync_config.get('modified_since'):
                modified_since = datetime.fromisoformat(sync_config['modified_since'])
                files_to_process = [
                    f for f in files_to_process
                    if datetime.fromtimestamp(os.path.getmtime(f)) > modified_since
                ]

            for file_path in files_to_process:
                try:
                    file_ext = os.path.splitext(file_path)[1].lower()

                    if file_ext in ['.xlsx', '.xls']:
                        data = await self.process_excel_file(file_path, integration.settings)
                    elif file_ext == '.csv':
                        data = await self.process_csv_file(file_path, integration.settings)
                    elif file_ext in ['.xml']:
                        data = await self.process_xml_file(file_path, integration.settings)
                    elif file_ext == '.json':
                        data = await self.process_json_file(file_path, integration.settings)
                    else:
                        self.logger.warning(f"Unsupported file type: {file_ext}")
                        continue

                    # Store or process data
                    file_name = os.path.basename(file_path)
                    await self.process_synced_data(integration.name, file_name, data)

                    records_synced += len(data) if isinstance(data, list) else 1

                except Exception as e:
                    self.logger.error(f"Error processing file {file_path}: {e}")

            return {'records_synced': records_synced}

        except Exception as e:
            self.logger.error(f"File sync error: {e}")
            raise

    async def sync_ftp_integration(self, integration: IntegrationConfig, sync_config: Dict[str, Any] = None) -> Dict[str, Any]:
        """Synchronize with FTP integration"""
        if not FTP_SSH_AVAILABLE:
            raise RuntimeError("FTP/SSH libraries not available")

        records_synced = 0

        try:
            # Setup FTP connection
            ftp_key = f"{integration.name}_ftp"
            if ftp_key not in self.ftp_connections:
                ftp = ftplib.FTP()
                ftp.connect(integration.endpoint.replace('ftp://', ''), 21)
                ftp.login(
                    integration.credentials.get('username', ''),
                    integration.credentials.get('password', '')
                )
                if integration.settings.get('passive_mode', True):
                    ftp.set_pasv(True)
                self.ftp_connections[ftp_key] = ftp

            ftp = self.ftp_connections[ftp_key]

            # List files
            files = ftp.nlst()

            # Filter files
            file_patterns = integration.settings.get('file_patterns', ['*'])
            download_path = integration.settings.get('download_path', '/tmp/')

            for file_name in files:
                try:
                    # Check if file matches patterns
                    import fnmatch
                    if not any(fnmatch.fnmatch(file_name, pattern) for pattern in file_patterns):
                        continue

                    # Download file
                    local_path = os.path.join(download_path, file_name)
                    with open(local_path, 'wb') as local_file:
                        ftp.retrbinary(f'RETR {file_name}', local_file.write)

                    # Process downloaded file
                    file_ext = os.path.splitext(file_name)[1].lower()
                    if file_ext in ['.xlsx', '.xls']:
                        data = await self.process_excel_file(local_path, integration.settings)
                    elif file_ext == '.csv':
                        data = await self.process_csv_file(local_path, integration.settings)
                    else:
                        continue

                    # Store or process data
                    await self.process_synced_data(integration.name, file_name, data)

                    records_synced += len(data) if isinstance(data, list) else 1

                    # Clean up
                    os.remove(local_path)

                except Exception as e:
                    self.logger.error(f"Error processing FTP file {file_name}: {e}")

            return {'records_synced': records_synced}

        except Exception as e:
            self.logger.error(f"FTP sync error: {e}")
            raise

    async def sync_ldap_integration(self, integration: IntegrationConfig, sync_config: Dict[str, Any] = None) -> Dict[str, Any]:
        """Synchronize with LDAP integration"""
        if not LDAP_AVAILABLE:
            raise RuntimeError("LDAP library not available")

        records_synced = 0

        try:
            # Setup LDAP connection
            server = ldap3.Server(integration.endpoint)
            conn = ldap3.Connection(
                server,
                user=integration.credentials.get('username', ''),
                password=integration.credentials.get('password', ''),
                auto_bind=True
            )

            # Search for users
            base_dn = integration.settings.get('base_dn', '')
            user_filter = integration.settings.get('user_filter', '(objectClass=person)')

            conn.search(base_dn, user_filter, attributes=['*'])

            users_data = []
            for entry in conn.entries:
                user_data = {
                    'dn': str(entry.entry_dn),
                    'attributes': dict(entry.entry_attributes_as_dict)
                }
                users_data.append(user_data)

            # Store or process data
            await self.process_synced_data(integration.name, 'users', users_data)

            records_synced = len(users_data)

            conn.unbind()

            return {'records_synced': records_synced}

        except Exception as e:
            self.logger.error(f"LDAP sync error: {e}")
            raise

    async def process_excel_file(self, file_path: str, settings: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Process Excel file"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not available for Excel processing")

        data = []

        try:
            workbook = load_workbook(file_path, read_only=True)

            # Auto-detect sheets or use specified ones
            if settings.get('auto_detect_sheets', True):
                sheets_to_process = workbook.sheetnames
            else:
                sheets_to_process = settings.get('sheets', [workbook.active.title])

            for sheet_name in sheets_to_process:
                sheet = workbook[sheet_name]
                header_row = settings.get('header_row', 1)

                # Get headers
                headers = []
                for cell in sheet[header_row]:
                    headers.append(cell.value if cell.value else f"col_{len(headers)}")

                # Read data rows
                for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        row_data = {}
                        for i, value in enumerate(row):
                            if i < len(headers):
                                row_data[headers[i]] = value
                        row_data['_sheet'] = sheet_name
                        data.append(row_data)

            workbook.close()

        except Exception as e:
            self.logger.error(f"Error processing Excel file {file_path}: {e}")
            raise

        return data

    async def process_csv_file(self, file_path: str, settings: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Process CSV file"""
        import csv
        import chardet

        data = []

        try:
            # Detect encoding
            with open(file_path, 'rb') as f:
                raw_data = f.read()
                result = chardet.detect(raw_data)
                encoding = result['encoding'] or 'utf-8'

            # Read CSV
            with open(file_path, 'r', encoding=encoding, newline='') as f:
                # Auto-detect delimiter
                sample = f.read(1024)
                f.seek(0)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter

                reader = csv.DictReader(f, delimiter=delimiter)
                for row in reader:
                    data.append(dict(row))

        except Exception as e:
            self.logger.error(f"Error processing CSV file {file_path}: {e}")
            raise

        return data

    async def process_xml_file(self, file_path: str, settings: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Process XML file"""
        data = []

        try:
            tree = ET.parse(file_path)
            root = tree.getroot()

            # Convert XML to dictionary
            def xml_to_dict(element):
                result = {}
                for child in element:
                    if len(child) == 0:
                        result[child.tag] = child.text
                    else:
                        result[child.tag] = xml_to_dict(child)
                return result

            if len(root) > 0:
                if len(root[0]) > 0:  # Multiple records
                    for child in root:
                        data.append(xml_to_dict(child))
                else:  # Single record
                    data.append(xml_to_dict(root))

        except Exception as e:
            self.logger.error(f"Error processing XML file {file_path}: {e}")
            raise

        return data

    async def process_json_file(self, file_path: str, settings: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Process JSON file"""
        data = []

        try:
            if AIOHTTP_AVAILABLE:
                async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
                    content = await f.read()
            else:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()

            json_data = json.loads(content)

            if isinstance(json_data, list):
                data = json_data
            else:
                data = [json_data]

        except Exception as e:
            self.logger.error(f"Error processing JSON file {file_path}: {e}")
            raise

        return data

    def parse_xml_data(self, xml_text: str) -> Dict[str, Any]:
        """Parse XML response data"""
        try:
            root = ET.fromstring(xml_text)

            def xml_to_dict(element):
                result = {}
                for child in element:
                    if len(child) == 0:
                        result[child.tag] = child.text
                    else:
                        result[child.tag] = xml_to_dict(child)
                return result

            return xml_to_dict(root)

        except Exception as e:
            self.logger.error(f"Error parsing XML data: {e}")
            return {'raw_xml': xml_text}

    async def setup_api_auth(self, integration: IntegrationConfig) -> Dict[str, str]:
        """Setup API authentication headers"""
        headers = {}

        auth_type = integration.settings.get('auth_type', 'basic')

        if auth_type == 'basic':
            username = integration.credentials.get('username', '')
            password = integration.credentials.get('password', '')
            credentials = base64.b64encode(f"{username}:{password}".encode()).decode()
            headers['Authorization'] = f'Basic {credentials}'

        elif auth_type == 'api_key':
            api_key = integration.credentials.get('api_key', '')
            key_header = integration.settings.get('api_key_header', 'X-API-Key')
            headers[key_header] = api_key

        elif auth_type == 'bearer':
            token = integration.credentials.get('token', '')
            headers['Authorization'] = f'Bearer {token}'

        elif auth_type == 'oauth2':
            # OAuth2 token exchange
            token = await self.get_oauth2_token(integration)
            headers['Authorization'] = f'Bearer {token}'

        # Add common headers
        headers.update({
            'Content-Type': 'application/json',
            'User-Agent': 'DAP-ThirdPartyIntegrator/1.0'
        })

        return headers

    async def get_oauth2_token(self, integration: IntegrationConfig) -> str:
        """Get OAuth2 access token"""
        # This is a simplified OAuth2 implementation
        # In production, use proper OAuth2 library
        token_url = integration.settings.get('token_url', '')
        client_id = integration.credentials.get('client_id', '')
        client_secret = integration.credentials.get('client_secret', '')

        if not all([token_url, client_id, client_secret]):
            raise ValueError("OAuth2 credentials incomplete")

        data = {
            'grant_type': 'client_credentials',
            'client_id': client_id,
            'client_secret': client_secret
        }

        async with self.http_session.post(token_url, data=data) as response:
            if response.status == 200:
                token_data = await response.json()
                return token_data.get('access_token', '')
            else:
                raise RuntimeError(f"OAuth2 token request failed: {response.status}")

    def build_connection_string(self, integration: IntegrationConfig) -> str:
        """Build database connection string"""
        endpoint = integration.endpoint
        username = integration.credentials.get('username', '')
        password = integration.credentials.get('password', '')

        if username and password:
            # Replace placeholder with actual credentials
            if '://' in endpoint:
                protocol, rest = endpoint.split('://', 1)
                return f"{protocol}://{username}:{password}@{rest}"
            else:
                return f"mssql://{username}:{password}@{endpoint}"

        return endpoint

    def apply_transformation(self, value: Any, transformation: str) -> Any:
        """Apply data transformation"""
        try:
            if transformation == 'upper':
                return str(value).upper() if value else value
            elif transformation == 'lower':
                return str(value).lower() if value else value
            elif transformation == 'strip':
                return str(value).strip() if value else value
            elif transformation == 'date':
                if isinstance(value, str):
                    from dateutil import parser
                    return parser.parse(value).isoformat()
                return value
            elif transformation.startswith('format:'):
                format_string = transformation.split(':', 1)[1]
                return format_string.format(value)
            else:
                return value
        except Exception as e:
            self.logger.warning(f"Transformation failed: {e}")
            return value

    async def process_synced_data(self, integration_name: str, data_type: str, data: List[Dict[str, Any]]):
        """Process synced data - override in subclass"""
        # This is where you would save the data to your local database
        # or pass it to the data processing pipeline
        self.logger.info(f"Processing {len(data)} records from {integration_name}.{data_type}")

        # Example: Save to file for testing
        import os
        output_dir = "sync_output"
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{integration_name}_{data_type}_{timestamp}.json"
        filepath = os.path.join(output_dir, filename)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def add_integration(self, name: str, config: IntegrationConfig):
        """Add new integration configuration"""
        self.integrations[name] = config
        self.logger.info(f"Added integration: {name}")

    def remove_integration(self, name: str):
        """Remove integration configuration"""
        if name in self.integrations:
            del self.integrations[name]
            if name in self.sync_statuses:
                del self.sync_statuses[name]
            if name in self.sync_schedules:
                del self.sync_schedules[name]
            self.logger.info(f"Removed integration: {name}")

    def schedule_sync(self, integration_name: str, interval: int, enabled: bool = True):
        """Schedule automatic synchronization"""
        self.sync_schedules[integration_name] = {
            'interval': interval,
            'enabled': enabled,
            'last_sync': None
        }
        self.logger.info(f"Scheduled sync for {integration_name} every {interval} seconds")

    def get_integration_status(self) -> Dict[str, Any]:
        """Get integration status"""
        return {
            'total_integrations': len(self.integrations),
            'active_syncs': len(self.active_syncs),
            'sync_statuses': {name: asdict(status) for name, status in self.sync_statuses.items()},
            'scheduled_syncs': len(self.sync_schedules),
            'scheduler_running': self.scheduler_running
        }

# Test function
async def test_third_party_integrator():
    """Test the third-party integrator functionality"""
    print("Testing Third-Party Integrator...")

    integrator = ThirdPartyIntegrator()
    print(f"✓ Integrator initialized: {len(integrator.integrations)} integrations")

    # Test integration config
    test_config = IntegrationConfig(
        name='Test Integration',
        type='api',
        endpoint='https://api.example.com',
        credentials={'api_key': 'test'},
        settings={'timeout': 30}
    )

    integrator.add_integration('test', test_config)
    print(f"✓ Integration added: {test_config.name}")

    print("✓ Third-Party Integrator test completed")

if __name__ == "__main__":
    import asyncio

    # Run test
    asyncio.run(test_third_party_integrator())